"""
Step 4 -- the output workbook writer.

Four tabs, reconciliation first (the packing slip goes on top of the
box): Reconciliation, Final Report, Completeness, Data Quality.

Read-only in, new file out -- this module never opens a source workbook
and never writes into data/. Deterministic: given the same inputs, the
only thing that varies between runs is the run timestamp, which this
module threads through once (workbook properties and the provenance
block both read from the same value) rather than letting openpyxl and
the provenance cell each take an independent, slightly-different
`datetime.now()`.
"""

import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from validator.checks import overall_status

DOC_CREATOR = "Substation Report Validator"
DOC_COMPANY = ""

BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
STATUS_FILL = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "CHECK": PatternFill("solid", fgColor="FFC7CE"),
    "NOT RUN": PatternFill("solid", fgColor="FFEB9C"),
}

_RULE_RE = re.compile(r"^R(\d+)([a-z]*)$")


def _rule_sort_key(rule):
    m = _RULE_RE.match(rule)
    if not m:
        return (999, rule)
    return (int(m.group(1)), m.group(2))


def _autosize(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


R0_MESSAGE = (
    "Governing rule: CHECK if any rule below is CHECK; otherwise NOT RUN "
    "if any rule below is NOT RUN; otherwise PASS. A NOT RUN rule is never "
    "rolled up as PASS -- V4 (R18-R20) is the one check a human can let "
    "lapse by leaving the G&T figure blank, and this status row must show "
    "that rather than hide it."
)


def build_reconciliation_rows(check_rows):
    """R0 + every check row, sorted by rule number -- the exact rows and
    order written to the Reconciliation tab's rule table, so console
    output built from this can never drift from the tab itself."""
    status = overall_status(check_rows)
    all_rows = [("R0", status, R0_MESSAGE)] + [(r.rule, r.result, r.message) for r in check_rows]
    all_rows.sort(key=lambda row: _rule_sort_key(row[0]))
    return status, all_rows


def _write_reconciliation_tab(ws, check_rows, config, interval_report, hourly_report, app_version, run_timestamp):
    status, all_rows = build_reconciliation_rows(check_rows)

    ws.cell(row=1, column=1, value="STATUS:").font = Font(bold=True, size=14)
    status_cell = ws.cell(row=1, column=2, value=status)
    status_cell.font = Font(bold=True, size=14)
    status_cell.fill = STATUS_FILL[status]
    ws.cell(row=2, column=1, value=(
        "PASS / CHECK / NOT RUN, in words, never a code or a colour alone. "
        "One status, one place, one word (R0)."
    )).font = Font(italic=True, size=9)

    table_start = 4
    headers = ["Rule", "Result", "Message"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=table_start, column=col_idx, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL

    row_idx = table_start
    for rule, result, message in all_rows:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=rule)
        result_cell = ws.cell(row=row_idx, column=2, value=result)
        result_cell.fill = STATUS_FILL[result]
        ws.cell(row=row_idx, column=3, value=message).alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, [8, 10, 120])

    prov_start = row_idx + 3
    ws.cell(row=prov_start, column=1, value="Provenance").font = Font(bold=True, size=12)
    tz = config["timezone"]
    provenance_lines = [
        f"Interval (15-minute) source file: {interval_report['path'].name}",
        f"Hourly source file: {hourly_report['path'].name}",
        f"Report month: {config['report_month']}",
        f"App version: {app_version}",
        f"Run timestamp: {run_timestamp.strftime('%Y-%m-%d %H:%M:%S')}",
        (
            f"Timezone assumption: timestamps are read as LOCAL time in "
            f"{tz['zone']}, not standard time year-round. On the fall-back "
            f"date, the export cannot distinguish the two 01:00 local hours "
            f"from each other; this app tolerates that duplicate rather than "
            f"resolving it (R15a)."
        ),
    ]
    for i, line in enumerate(provenance_lines, start=1):
        ws.cell(row=prov_start + i, column=1, value=line)


def _write_final_report_tab(ws, config, final_report_data):
    headers = config["final_report_tab"]["headers"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL

    first_data_row = 2
    rows = final_report_data["rows"]
    for i, row_data in enumerate(rows):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=row_data["generic_id"])
        ws.cell(row=r, column=2, value=row_data["meter_label"])
        ws.cell(row=r, column=3, value=row_data["c_hourly_kwh"])
        ws.cell(row=r, column=4, value=row_data["d_interval_kwh"])
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)")
        ws.cell(row=r, column=6).number_format = "0.0%"

    last_data_row = first_data_row + len(rows) - 1
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})").font = BOLD
    ws.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})").font = BOLD
    ws.cell(row=total_row, column=5, value=f"=C{total_row}-D{total_row}").font = BOLD
    total_f = ws.cell(row=total_row, column=6, value=f"=IFERROR(E{total_row}/C{total_row},0)")
    total_f.font = BOLD
    total_f.number_format = "0.0%"

    control_row_1 = total_row + 2
    control_row_2 = total_row + 3
    ws.cell(row=control_row_1, column=1,
            value="Total System by Substation Meter (control total, from the hourly file's own Total/System column)")
    ws.cell(row=control_row_1, column=3, value=final_report_data["control_hourly_total_system"])
    ws.cell(row=control_row_2, column=1,
            value="Total System by Substation (control total, from the 15-minute file's own Total/System column)")
    ws.cell(row=control_row_2, column=4, value=final_report_data["control_interval_total_system"])

    _autosize(ws, [18, 20, 20, 20, 20, 14])


def _write_completeness_tab(ws, roster, hourly_grid, interval_grid):
    row_idx = 1

    def write_grid(title, grid, expected_label):
        nonlocal row_idx
        ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=12)
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=(
            f"Reading count per substation per date. Blank means zero readings "
            f"for the entire date -- an absent day is visible as blank, not 0. "
            f"{expected_label}"
        )).font = Font(italic=True, size=9)
        row_idx += 1

        header_row = row_idx
        ws.cell(row=header_row, column=1, value="Date").font = BOLD
        ws.cell(row=header_row, column=1).fill = HEADER_FILL
        for col_idx, generic_id in enumerate(roster, start=2):
            c = ws.cell(row=header_row, column=col_idx, value=generic_id)
            c.font = BOLD
            c.fill = HEADER_FILL

        for date_str, counts in grid.items():
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=date_str)
            for col_idx, generic_id in enumerate(roster, start=2):
                count = counts.get(generic_id, 0)
                ws.cell(row=row_idx, column=col_idx, value=(count if count else None))
        row_idx += 3

    write_grid(
        "Hourly file completeness",
        hourly_grid,
        "In a clean month with no DST transition, every cell reads 24.",
    )
    write_grid(
        "15-minute file completeness",
        interval_grid,
        "In a clean month with no DST transition, every cell reads 96.",
    )

    _autosize(ws, [12] + [8] * len(roster))


def _write_data_quality_tab(ws, exception_flags_config, interval_counts, hourly_counts, notes):
    headers = ["Exception Flag", "Interval File Count", "Hourly File Count", "Note"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL

    row_idx = 1
    for flag_name in exception_flags_config:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=flag_name)
        ws.cell(row=row_idx, column=2, value=interval_counts.get(flag_name, 0))
        ws.cell(row=row_idx, column=3, value=hourly_counts.get(flag_name, 0))
        note = notes.get(flag_name, "")
        if note:
            ws.cell(row=row_idx, column=4, value=note).alignment = Alignment(wrap_text=True)

    _autosize(ws, [24, 18, 18, 70])


def build_workbook(
    interval_report, hourly_report, config, crosswalk_rows,
    check_rows, final_report_data, hourly_grid, interval_grid,
    interval_exception_counts, hourly_exception_counts, exception_notes,
    app_version, out_path,
):
    run_timestamp = datetime.now()

    wb = Workbook()
    wb.properties.creator = DOC_CREATOR
    wb.properties.lastModifiedBy = DOC_CREATOR
    wb.properties.company = DOC_COMPANY
    wb.properties.title = f"Substation Report -- {config['report_month']}"
    wb.properties.created = run_timestamp
    wb.properties.modified = run_timestamp

    ws_recon = wb.active
    ws_recon.title = "Reconciliation"
    _write_reconciliation_tab(
        ws_recon, check_rows, config, interval_report, hourly_report, app_version, run_timestamp
    )

    ws_final = wb.create_sheet("Final Report")
    _write_final_report_tab(ws_final, config, final_report_data)

    ws_complete = wb.create_sheet("Completeness")
    roster = [row["generic_id"] for row in crosswalk_rows]
    _write_completeness_tab(ws_complete, roster, hourly_grid, interval_grid)

    ws_quality = wb.create_sheet("Data Quality")
    _write_data_quality_tab(
        ws_quality, config["exception_flags"], interval_exception_counts, hourly_exception_counts, exception_notes
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
