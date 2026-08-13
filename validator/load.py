"""
Step 1 loader.

Opens the two source workbooks read-only and reports what is actually in
them: sheet names, header text, row/column counts, parsed timestamps, the
unit row, the total row, and how each column lines up against
crosswalk.csv. It runs no R1-R20 rules and makes no PASS/CHECK judgement --
that starts in Step 2.

Timestamps are still parsed with an explicit format string and the parse
still fails loudly on a row that doesn't match it (constraint 1). That is
baseline correct parsing, not one of the deferred validation rules.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class LoadError(Exception):
    """Fail-loud errors for file discovery and timestamp parsing."""


def find_source_file(config, pattern_key, repo_root):
    """Locate the one file matching config[pattern_key] under
    config['source_folder']. Never guesses which file to use -- zero or
    more-than-one matches is an error, named explicitly."""
    source_folder = repo_root / config["source_folder"]
    pattern = config[pattern_key]
    if not source_folder.exists():
        raise LoadError(f"source_folder not found: {source_folder}")

    matches = sorted(source_folder.glob(pattern))
    if not matches:
        raise LoadError(
            f"No file matched pattern {pattern!r} ({pattern_key}) in {source_folder}"
        )
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        raise LoadError(
            f"Multiple files matched pattern {pattern!r} ({pattern_key}) in "
            f"{source_folder}: {names}. Expected exactly one."
        )
    return matches[0]


def _cell_display(value):
    if value is None:
        return "(blank)"
    return repr(value)


def discover_workbook(path, role, config, crosswalk_rows, position_field):
    """Open one workbook and build a plain-data report of what's in it.

    role: "interval" or "hourly" -- selects which crosswalk position field
    and timestamp handling (text vs. native datetime) applies.
    position_field: "interval_position" or "hourly_position".
    """
    layout = config["layout"]
    header_row_idx = layout["header_row"]
    unit_row_idx = layout["unit_description_row"]
    first_data_row_idx = layout["first_data_row"]
    last_row_is_total_expected = layout["last_row_is_total"]
    leading_ts_cols = layout["leading_timestamp_columns"]
    primary_ts_col = layout["timestamp_column_index"]
    timestamps_are_text = layout["interval_timestamps_are_text"] if role == "interval" \
        else not layout["hourly_timestamps_are_datetime"]
    ts_format = layout["interval_timestamp_format"]

    # read_only=True has pathologically slow random cell() access in
    # openpyxl (each call can rescan from the start of the sheet); these
    # files are small (a few hundred KB), so load them fully instead.
    wb = load_workbook(path, data_only=True)
    sheet_names = list(wb.sheetnames)
    ws = wb[sheet_names[0]]

    max_column = ws.max_column
    max_row = ws.max_row

    header_cells = [ws.cell(row=header_row_idx, column=c).value for c in range(1, max_column + 1)]
    unit_row_cells = [ws.cell(row=unit_row_idx, column=c).value for c in range(1, max_column + 1)]

    last_row_label = ws.cell(row=max_row, column=1).value
    total_row_detected = (
        last_row_is_total_expected
        and isinstance(last_row_label, str)
        and "total" in last_row_label.lower()
    )
    last_data_row_idx = (max_row - 1) if total_row_detected else max_row
    data_row_count = last_data_row_idx - first_data_row_idx + 1

    total_col_idx = max_column  # config: columns.total_system_column == "last"
    location_col_start = leading_ts_cols + 1
    location_col_end = total_col_idx - 1
    location_count = location_col_end - location_col_start + 1

    # --- crosswalk match, per location column ---
    # Computed before the data-row pass below: it depends only on the
    # header row and crosswalk.csv, and the data pass needs column_matches
    # already built so it can key each cell's value by generic_id as it goes.
    crosswalk_by_position = {row[position_field]: row for row in crosswalk_rows}
    header_field = "expected_interval_header" if role == "interval" else "expected_hourly_header"

    column_matches = []
    for col_idx in range(location_col_start, location_col_end + 1):
        position = col_idx - leading_ts_cols
        header_text = header_cells[col_idx - 1]
        cw_row = crosswalk_by_position.get(position)
        if cw_row is None:
            column_matches.append(
                {
                    "column": col_idx,
                    "letter": get_column_letter(col_idx),
                    "header": header_text,
                    "position": position,
                    "status": "UNMAPPED - no crosswalk entry at this position",
                }
            )
            continue
        expected_header = cw_row[header_field]
        if not expected_header:
            status = (
                f"position matched -> {cw_row['generic_id']} "
                f"(crosswalk {header_field} is blank; header text not verified)"
            )
        elif expected_header.strip() == (header_text or "").strip():
            status = f"position matched -> {cw_row['generic_id']} (header text verified)"
        else:
            status = (
                f"position matched -> {cw_row['generic_id']}, but header text differs: "
                f"crosswalk expects {expected_header!r}, file has {header_text!r}"
            )
        column_matches.append(
            {
                "column": col_idx,
                "letter": get_column_letter(col_idx),
                "header": header_text,
                "position": position,
                "generic_id": cw_row["generic_id"],
                "status": status,
            }
        )

    expected_positions = {row[position_field] for row in crosswalk_rows}
    found_positions = {m["position"] for m in column_matches}
    missing_positions = sorted(expected_positions - found_positions)

    # --- timestamps, and (same pass) each matched location column's raw
    # cell value keyed by generic_id and row timestamp. V3 (R11-R17) reads
    # readings_by_generic_id to judge completeness; a blank cell there is
    # "no reading", not a parse error -- unlike the timestamp columns,
    # location cells are not required to be non-blank at this layer.
    ts_columns = list(range(1, leading_ts_cols + 1))
    parsed_by_col = {c: [] for c in ts_columns}
    matched_location_columns = [m for m in column_matches if "generic_id" in m]
    readings_by_generic_id = {m["generic_id"]: [] for m in matched_location_columns}

    for row_idx in range(first_data_row_idx, last_data_row_idx + 1):
        for c in ts_columns:
            raw = ws.cell(row=row_idx, column=c).value
            if timestamps_are_text:
                if not isinstance(raw, str):
                    raise LoadError(
                        f"{path.name}: row {row_idx}, column {get_column_letter(c)}: "
                        f"expected text timestamp, found {type(raw).__name__} {raw!r}"
                    )
                try:
                    parsed = datetime.strptime(raw, ts_format)
                except ValueError as e:
                    raise LoadError(
                        f"{path.name}: row {row_idx}, column {get_column_letter(c)}: "
                        f"timestamp {raw!r} does not match expected format {ts_format!r} ({e})"
                    ) from e
            else:
                if not isinstance(raw, datetime):
                    raise LoadError(
                        f"{path.name}: row {row_idx}, column {get_column_letter(c)}: "
                        f"expected a native Excel datetime, found {type(raw).__name__} {raw!r}"
                    )
                parsed = raw
            parsed_by_col[c].append(parsed)

        row_ts = parsed_by_col[primary_ts_col][-1]
        for m in matched_location_columns:
            value = ws.cell(row=row_idx, column=m["column"]).value
            readings_by_generic_id[m["generic_id"]].append((row_ts, value))

    primary_first = parsed_by_col[primary_ts_col][0]
    primary_last = parsed_by_col[primary_ts_col][-1]

    wb.close()

    return {
        "role": role,
        "path": path,
        "sheet_names": sheet_names,
        "sheet_used": sheet_names[0],
        "max_column": max_column,
        "max_row": max_row,
        "header_row_idx": header_row_idx,
        "unit_row_idx": unit_row_idx,
        "headers": header_cells,
        "unit_row_cells": unit_row_cells,
        "first_data_row_idx": first_data_row_idx,
        "last_data_row_idx": last_data_row_idx,
        "data_row_count": data_row_count,
        "last_row_label": last_row_label,
        "total_row_detected": total_row_detected,
        "total_col_idx": total_col_idx,
        "location_col_start": location_col_start,
        "location_col_end": location_col_end,
        "location_count": location_count,
        "ts_columns": ts_columns,
        "primary_ts_col": primary_ts_col,
        "timestamps_are_text": timestamps_are_text,
        "ts_format": ts_format if timestamps_are_text else None,
        "primary_first_ts": primary_first,
        "primary_last_ts": primary_last,
        "primary_ts_values": parsed_by_col[primary_ts_col],
        "column_matches": column_matches,
        "missing_crosswalk_positions": missing_positions,
        "readings_by_generic_id": readings_by_generic_id,
    }


def print_report(report):
    p = report["path"]
    print("=" * 70)
    print(f"FILE ({report['role']}): {p.name}")
    print("=" * 70)
    print(f"Path: {p}")
    print(f"Sheet names found: {report['sheet_names']}")
    print(f"Using sheet: {report['sheet_used']!r}")
    print()

    print(f"Header row (row {report['header_row_idx']}), "
          f"{report['max_column']} columns, in file order:")
    for i, h in enumerate(report["headers"], start=1):
        print(f"  {i:2d} ({get_column_letter(i)}): {_cell_display(h)}")
    print()

    print(f"Row {report['unit_row_idx']} (unit/description row) -- skipped as non-data:")
    for i, u in enumerate(report["unit_row_cells"], start=1):
        print(f"  {i:2d} ({get_column_letter(i)}): {_cell_display(u)}")
    print()

    print(f"Data rows: {report['first_data_row_idx']} through "
          f"{report['last_data_row_idx']} ({report['data_row_count']} rows)")
    last_idx = report["max_row"]
    if report["total_row_detected"]:
        print(f"Last row (row {last_idx}): label={report['last_row_label']!r} "
              f"-> identified as Total row")
    else:
        print(f"Last row (row {last_idx}): label={report['last_row_label']!r} "
              f"-> NOT identified as a Total row (no validation enforced yet)")
    print()

    print(f"Timestamp columns: {report['ts_columns']} "
          f"(primary = column {report['primary_ts_col']})")
    if report["timestamps_are_text"]:
        print(f"  Stored as TEXT. Parse format used: {report['ts_format']!r}")
    else:
        print("  Stored as native Excel datetime values (no text parsing).")
    print(f"  All {report['data_row_count']} rows parsed successfully in every timestamp column.")
    print(f"  First timestamp (primary column): {report['primary_first_ts']}")
    print(f"  Last timestamp (primary column):  {report['primary_last_ts']}")
    print()

    print(f"Crosswalk match ({report['location_count']} location columns, "
          f"columns {report['location_col_start']}-{report['location_col_end']}):")
    for m in report["column_matches"]:
        print(f"  {m['column']:2d} ({m['letter']}) {m['header']!r:>14} "
              f"-> position {m['position']:2d}: {m['status']}")
    if report["missing_crosswalk_positions"]:
        print(f"  Crosswalk positions with NO matching column in this file: "
              f"{report['missing_crosswalk_positions']}")
    print(f"  {report['total_col_idx']:2d} "
          f"({get_column_letter(report['total_col_idx'])}) "
          f"{report['headers'][report['total_col_idx'] - 1]!r:>14} "
          f"-> Total/System control column (excluded from crosswalk by design)")
    print()
