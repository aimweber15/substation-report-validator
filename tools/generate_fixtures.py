#!/usr/bin/env python
"""
Builds the ten test fixtures into tests/fixtures/<name>/, each with its
own tiny config.json.

Every fixture but F8/F8b is the clean July sample with exactly ONE
change applied via openpyxl -- copy first, then mutate, so the delta is
literal and inspectable, never a from-scratch rebuild that could drift.
F8 and F8b are the 25-hour November 2026 case, built the same way the
July sample is (tools/generate_sample_data.py's write_workbook), but
with the interval/hourly timestamp lists sourced from
validator.stamps.expected_stamps()'s timezone-aware walk instead of the
naive walk that function uses for non-DST months -- that walk is what
puts 11/01/2026 01:00 into the list twice, which the plain walk cannot.

No real substation names, meter IDs, or readings anywhere -- every value
here is either copied from the already-synthetic sample or a plain
placeholder number.
"""

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "tools"))
sys.path.insert(0, str(REPO_ROOT))

from generate_sample_data import (  # noqa: E402
    HOURLY_NOISE_SEED, INTERVAL_NOISE_SEED, INTERVAL_TIMESTAMP_FORMAT,
    build_location_meta, load_crosswalk, write_workbook,
)
from validator.stamps import expected_stamps  # noqa: E402

FIXTURES_DIR = REPO_ROOT / "tests" / "fixtures"
JULY_INTERVAL_SRC = REPO_ROOT / "data" / "coop_15min_2026-07_SAMPLE.xlsx"
JULY_HOURLY_SRC = REPO_ROOT / "data" / "coop_hourly_2026-07_SAMPLE.xlsx"

INTERVAL_FIRST_DATA_ROW = 3
INTERVAL_LAST_DATA_ROW = 2978
INTERVAL_TOTAL_ROW = 2979
HOURLY_FIRST_DATA_ROW = 3
HOURLY_LAST_DATA_ROW = 746
HOURLY_TOTAL_ROW = 747
MAX_COL = 19
TOTAL_COL = 19


def _fresh_dir(name):
    d = FIXTURES_DIR / name
    if d.exists():
        shutil.rmtree(d)
    d.mkdir(parents=True)
    return d


def _write_config(dest_dir, report_month):
    """Tiny config: every key load_config() actually requires, nothing a
    fixture doesn't need (no final_report_tab/exception_flags/scope --
    those feed Step 4's workbook, not the V1-V3 checks fixtures exercise).
    source_folder is relative to REPO_ROOT, same as the real config.json.
    """
    rel_source = (dest_dir.relative_to(REPO_ROOT)).as_posix()
    config = {
        "report_month": report_month,
        "source_folder": f"./{rel_source}",
        "interval_file_pattern": "*15*_FIXTURE.xlsx",
        "hourly_file_pattern": "*hourly*_FIXTURE.xlsx",
        "layout": {
            "structure": "wide",
            "header_row": 1,
            "unit_description_row": 2,
            "first_data_row": 3,
            "last_row_is_total": True,
            "leading_timestamp_columns": 2,
            "timestamp_column_index": 1,
            "timestamp_column_meaning": "interval_start",
            "interval_timestamps_are_text": True,
            "interval_timestamp_format": "%m/%d/%Y %H:%M",
            "hourly_timestamps_are_datetime": True,
        },
        "columns": {
            "individual_count": 16,
            "total_system_column": "last",
            "total_system_is_control_only": True,
        },
        "crosswalk_file": "./crosswalk.csv",
        "match_by": "position_with_header_verification",
        "units": {
            "interval_value_unit": "kWh",
            "hourly_value_unit": "kWh",
            "aggregation_method": "sum",
        },
        "expected_interval_minutes": 15,
        "hourly_interval_minutes": 60,
        "timezone": {
            "mode": "local",
            "zone": "America/Chicago",
            "expected_rows_per_day": "computed",
        },
        "reconciliation": {
            "gt_figure": None,
            "tolerance_pct": 2.0,
            "tolerance_basis": "energy_only",
            "tie_out_is_external": True,
            "difference_formula": "meter_kwh - substation_kwh",
            "loss_percent_denominator": "meter_kwh",
            "meter_kwh_source": "hourly_file",
            "substation_kwh_source": "interval_file",
            "emit_status_column": False,
        },
    }
    config_path = dest_dir / "config.json"
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)
        f.write("\n")
    return config_path


def _copy_clean_july(dest_dir):
    interval_dst = dest_dir / "coop_15min_2026-07_FIXTURE.xlsx"
    hourly_dst = dest_dir / "coop_hourly_2026-07_FIXTURE.xlsx"
    shutil.copy(JULY_INTERVAL_SRC, interval_dst)
    shutil.copy(JULY_HOURLY_SRC, hourly_dst)
    return interval_dst, hourly_dst


def _interval_day_row_range(ws, date_prefix):
    rows = [
        r for r in range(INTERVAL_FIRST_DATA_ROW, INTERVAL_LAST_DATA_ROW + 1)
        if ws.cell(row=r, column=1).value.startswith(date_prefix)
    ]
    if not rows:
        raise ValueError(f"no interval rows found for date prefix {date_prefix!r}")
    return min(rows), max(rows), len(rows)


def _find_total_row(ws, max_row_hint):
    r = max_row_hint
    while ws.cell(row=r, column=1).value != "Total":
        r -= 1
        if r < 1:
            raise ValueError("Total row not found")
    return r


# ---------------------------------------------------------------- F0 ----

def build_f0():
    """Clean July, unmodified. Must trip nothing."""
    dest = _fresh_dir("f0_clean_july")
    _copy_clean_july(dest)
    _write_config(dest, "2026-07")
    print(f"F0  wrote {dest}")


# ---------------------------------------------------------------- F8 ----

def _november_timestamps(step_minutes):
    stamps, _ = expected_stamps("2026-11", step_minutes, "America/Chicago")
    fmt = INTERVAL_TIMESTAMP_FORMAT
    return [datetime.strptime(s, fmt) for s in stamps]


def _build_november(dest_dir, interval_ts, hourly_ts, label):
    crosswalk_rows = load_crosswalk()
    meta = build_location_meta(crosswalk_rows)

    interval_order = sorted(crosswalk_rows, key=lambda r: r["interval_position"])
    hourly_order = sorted(crosswalk_rows, key=lambda r: r["hourly_position"])
    interval_ids = [r["generic_id"] for r in interval_order]
    hourly_ids = [r["generic_id"] for r in hourly_order]
    interval_headers = {gid: gid for gid in interval_ids}
    hourly_headers = {gid: meta[gid]["meter_label"] for gid in hourly_ids}

    interval_path = dest_dir / "coop_15min_2026-11_FIXTURE.xlsx"
    hourly_path = dest_dir / "coop_hourly_2026-11_FIXTURE.xlsx"

    interval_summary = write_workbook(
        interval_path, interval_ids, interval_headers, meta, interval_ts,
        15, timestamps_as_text=True, noise_seed=INTERVAL_NOISE_SEED,
    )
    hourly_summary = write_workbook(
        hourly_path, hourly_ids, hourly_headers, meta, hourly_ts,
        60, timestamps_as_text=False, noise_seed=HOURLY_NOISE_SEED,
    )
    _write_config(dest_dir, "2026-11")
    print(f"{label}  wrote {interval_path} ({interval_summary['rows']} rows), "
          f"{hourly_path} ({hourly_summary['rows']} rows)")


def build_f8():
    """November 2026, tz-aware: 25-hour Nov 1, 01:00 doubled. Must trip nothing."""
    dest = _fresh_dir("f8_november_clean")
    interval_ts = _november_timestamps(15)
    hourly_ts = _november_timestamps(60)
    _build_november(dest, interval_ts, hourly_ts, "F8 ")


def build_f8b():
    """Same November data, minus the SECOND 11/01 01:00 hourly occurrence
    (list.remove drops only the first match it finds, so remove it twice
    to guarantee the second/ambiguous occurrence is the one gone -- we
    want ONE surviving 01:00, not zero). Trips R14 (and, unavoidably,
    R12 -- see the module docstring in tools/run_fixtures.py)."""
    dest = _fresh_dir("f8b_november_25th_hour_missing")
    interval_ts = _november_timestamps(15)
    hourly_ts = _november_timestamps(60)

    ambiguous = datetime(2026, 11, 1, 1, 0)
    idx = [i for i, ts in enumerate(hourly_ts) if ts == ambiguous]
    assert len(idx) == 2, f"expected the fall-back hour twice, found {len(idx)}"
    del hourly_ts[idx[1]]  # drop the second occurrence only

    _build_november(dest, interval_ts, hourly_ts, "F8b")


# ---------------------------------------------------------------- F5 ----

def build_f5():
    """Interval file: 07/14 fully deleted, 07/21 fully duplicated (appended
    before the Total row). Row count is unchanged; the data is wrong.
    Trips R12 and R15 -- and, unavoidably, R14 on both affected dates
    (see tools/run_fixtures.py)."""
    dest = _fresh_dir("f5_missing_day_and_duplicated_day")
    interval_dst, _ = _copy_clean_july(dest)

    wb = load_workbook(interval_dst)
    ws = wb.active

    dup_start, dup_end, dup_count = _interval_day_row_range(ws, "07/21/2026")
    dup_values = [
        [ws.cell(row=r, column=c).value for c in range(1, MAX_COL + 1)]
        for r in range(dup_start, dup_end + 1)
    ]

    del_start, del_end, del_count = _interval_day_row_range(ws, "07/14/2026")
    ws.delete_rows(del_start, del_count)

    total_row = _find_total_row(ws, INTERVAL_TOTAL_ROW - del_count)
    ws.insert_rows(total_row, dup_count)
    for i, row_vals in enumerate(dup_values):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=total_row + i, column=c, value=val)

    wb.save(interval_dst)
    _write_config(dest, "2026-07")
    print(f"F5  {interval_dst.name}: deleted 07/14/2026 ({del_count} rows), "
          f"duplicated 07/21/2026 ({dup_count} rows) -- net row count unchanged")


# ---------------------------------------------------------------- F4 ----

def build_f4():
    """Hourly file: Sub D's (interval_position 4) 24 cells on 07/10 blanked.
    One substation, one date. Trips R12, R14, R17."""
    dest = _fresh_dir("f4_one_substation_one_date_removed")
    _, hourly_dst = _copy_clean_july(dest)

    wb = load_workbook(hourly_dst)
    ws = wb.active

    crosswalk_rows = load_crosswalk()
    target = next(r for r in crosswalk_rows if r["generic_id"] == "Sub D")
    col = 2 + target["hourly_position"]  # leading_timestamp_columns(2) + position

    count = 0
    for r in range(HOURLY_FIRST_DATA_ROW, HOURLY_LAST_DATA_ROW + 1):
        ts = ws.cell(row=r, column=1).value
        if ts.strftime("%m/%d") == "07/10":
            ws.cell(row=r, column=col).value = None  # ws.cell(..., value=None) is a no-op; must assign .value directly
            count += 1

    wb.save(hourly_dst)
    _write_config(dest, "2026-07")
    print(f"F4  {hourly_dst.name}: blanked Sub D's column on 07/10/2026 ({count} cells)")


# ---------------------------------------------------------------- F3 ----

def build_f3():
    """Interval file: one extra row at 08/01/2026 00:00, inserted before
    the Total row. Trips R7 and, by design, R9 alongside it (R9 is R7's
    own reporting requirement, not an independent trigger) -- and R10."""
    dest = _fresh_dir("f3_next_month_row")
    interval_dst, _ = _copy_clean_july(dest)

    wb = load_workbook(interval_dst)
    ws = wb.active

    total_row = _find_total_row(ws, INTERVAL_TOTAL_ROW)
    ws.insert_rows(total_row, 1)
    stamp_start = "08/01/2026 00:00"
    stamp_end = "08/01/2026 00:15"
    ws.cell(row=total_row, column=1, value=stamp_start)
    ws.cell(row=total_row, column=2, value=stamp_end)
    row_total = 0.0
    for c in range(3, MAX_COL):  # 16 location columns
        value = 50.0
        ws.cell(row=total_row, column=c, value=value)
        row_total += value
    ws.cell(row=total_row, column=MAX_COL, value=round(row_total, 1))

    wb.save(interval_dst)
    _write_config(dest, "2026-07")
    print(f"F3  {interval_dst.name}: inserted one row at {stamp_start}")


# ---------------------------------------------------------------- F1 ----

def build_f1():
    """Interval file: Sub C's header renamed to an arbitrary new string
    (not any other expected header). Trips R3 and, unavoidably, R4 --
    see tools/run_fixtures.py's module docstring for why a genuine
    rename cannot trip R3 alone once every crosswalk header is filled in."""
    dest = _fresh_dir("f1_header_renamed")
    interval_dst, _ = _copy_clean_july(dest)

    wb = load_workbook(interval_dst)
    ws = wb.active
    before = ws.cell(row=1, column=5).value  # column E = position 3 = Sub C
    ws.cell(row=1, column=5, value="Sub C RENAMED")
    wb.save(interval_dst)
    _write_config(dest, "2026-07")
    print(f"F1  {interval_dst.name}: header at column E renamed {before!r} -> 'Sub C RENAMED'")


# ---------------------------------------------------------------- F7 ----

def build_f7():
    """Interval file: Sub G's entire column blanked for the whole month
    (column stays, header stays -- only the data is gone). Trips R16,
    and unavoidably R12 and R14 across all 31 dates -- see
    tools/run_fixtures.py's module docstring."""
    dest = _fresh_dir("f7_substation_absent_all_month")
    interval_dst, _ = _copy_clean_july(dest)

    wb = load_workbook(interval_dst)
    ws = wb.active

    crosswalk_rows = load_crosswalk()
    target = next(r for r in crosswalk_rows if r["generic_id"] == "Sub G")
    col = 2 + target["interval_position"]

    count = 0
    for r in range(INTERVAL_FIRST_DATA_ROW, INTERVAL_LAST_DATA_ROW + 1):
        ws.cell(row=r, column=col).value = None  # ws.cell(..., value=None) is a no-op; must assign .value directly
        count += 1

    wb.save(interval_dst)
    _write_config(dest, "2026-07")
    print(f"F7  {interval_dst.name}: blanked Sub G's column entirely ({count} cells)")


# ---------------------------------------------------------------- F6 ----

def build_f6():
    """Hourly file: the last 4 hourly rows of 07/10 deleted outright (all
    16 substations lose those 4 hours simultaneously), leaving 20 rows
    that date instead of 24. Trips R14, and unavoidably R12 (those 4
    stamps are genuinely missing for every substation) -- see
    tools/run_fixtures.py's module docstring."""
    dest = _fresh_dir("f6_partial_day_20_of_24")
    _, hourly_dst = _copy_clean_july(dest)

    wb = load_workbook(hourly_dst)
    ws = wb.active

    rows_that_date = [
        r for r in range(HOURLY_FIRST_DATA_ROW, HOURLY_LAST_DATA_ROW + 1)
        if ws.cell(row=r, column=1).value.strftime("%m/%d") == "07/10"
    ]
    to_delete = sorted(rows_that_date)[-4:]  # last 4 hours of that date
    for r in reversed(to_delete):  # delete bottom-up so earlier indices stay valid
        ws.delete_rows(r, 1)

    wb.save(hourly_dst)
    _write_config(dest, "2026-07")
    print(f"F6  {hourly_dst.name}: deleted 4 rows from 07/10/2026 (20 of 24 remain)")


# ---------------------------------------------------------------- F2 ----

def build_f2():
    """Interval file: one new, unmapped location column inserted just
    before the Total/System column. Trips R4 only."""
    dest = _fresh_dir("f2_extra_unmapped_column")
    interval_dst, _ = _copy_clean_july(dest)

    wb = load_workbook(interval_dst)
    ws = wb.active

    new_col = TOTAL_COL  # insert before the current last (Total/System) column
    ws.insert_cols(new_col, 1)
    ws.cell(row=1, column=new_col, value="Sub Q (unmapped)")
    ws.cell(row=2, column=new_col, value="kWh")
    for r in range(INTERVAL_FIRST_DATA_ROW, INTERVAL_LAST_DATA_ROW + 1):
        ws.cell(row=r, column=new_col, value=10.0)
    total_row = _find_total_row(ws, INTERVAL_TOTAL_ROW)
    ws.cell(row=total_row, column=new_col, value=round(10.0 * (INTERVAL_LAST_DATA_ROW - INTERVAL_FIRST_DATA_ROW + 1), 1))

    wb.save(interval_dst)
    _write_config(dest, "2026-07")
    print(f"F2  {interval_dst.name}: inserted unmapped column at position {new_col}")


def main():
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    build_f0()
    build_f8()
    build_f5()
    build_f4()
    build_f3()
    build_f1()
    build_f7()
    build_f6()
    build_f2()
    build_f8b()
    print()
    print("All ten fixtures built.")


if __name__ == "__main__":
    main()
