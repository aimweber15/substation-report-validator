"""
Step 0 — Sample data generator for the Substation Report Validator.

Writes two fabricated Excel workbooks into data/: a 15-minute interval file
and an hourly file, shaped like the real cooperative's exports described in
config.json and crosswalk.csv but containing no real substation names, no
real meter numbers, and no real readings. Every number in these files is
synthetic. Document properties (creator/title/company) are set explicitly
in write_workbook() rather than left to openpyxl's defaults, so a
regenerated sample is clean by construction, not by scrubbing afterward.

Run:
    python tools/generate_sample_data.py

Deterministic: fixed random seeds below mean re-running this script
regenerates byte-for-byte identical data (aside from workbook metadata
openpyxl itself may stamp, e.g. calcChain/created time).
"""

import csv
import json
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = REPO_ROOT / "config.json"
CROSSWALK_PATH = REPO_ROOT / "crosswalk.csv"
DATA_DIR = REPO_ROOT / "data"

# Confirmed Aug 12 (see config.json "layout") and matched here: timestamps
# are interval START, column 2 is interval END, and this is the text
# format the sample data is written in.
INTERVAL_TIMESTAMP_FORMAT = "%m/%d/%Y %H:%M"
TIMESTAMP_COLUMN_MEANING = "interval_start (column 1); column 2 is interval_end"

SHEET_NAME = "Sheet1"
TOTAL_COLUMN_HEADER = "Total/System"
UNIT_LABEL = "kWh"

# Explicit, non-identifying document properties -- set on every workbook
# this script writes, so a regenerated sample never inherits a real name
# from the machine or account that ran the generator.
DOC_CREATOR = "Substation Report Validator sample data generator"
DOC_COMPANY = ""

# Daily load-curve anchors: hour of day -> fraction of that location's peak.
# Overnight low, morning ramp, midday shoulder, evening peak ~20:00.
ANCHOR_HOURS = [0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]
ANCHOR_VALS = [0.38, 0.33, 0.32, 0.42, 0.62, 0.72, 0.70, 0.68, 0.75, 0.92, 1.00, 0.70, 0.38]
WEEKEND_FACTOR = 0.85

# Solar dip: subtracted from net load at solar locations, bell-shaped around
# midday. Narrow enough that hourly sampling smooths it differently than
# 15-minute sampling does, so the two files do not reconcile exactly for
# Sub H, O, P -- a real difference, not a bug in the generator.
SOLAR_PEAK_HOUR = 13.0
SOLAR_SIGMA_HOURS = 2.6
SOLAR_PEAK_FRACTION = 0.55

CAPACITY_SEED = 2026
INTERVAL_NOISE_SEED = 20260701
HOURLY_NOISE_SEED = 20260702


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_crosswalk():
    rows = []
    with open(CROSSWALK_PATH, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(
                {
                    "generic_id": row["generic_id"].strip(),
                    "interval_position": int(row["interval_position"]),
                    "hourly_position": int(row["hourly_position"]),
                    "location_type": row["location_type"].strip(),
                }
            )
    return rows


def day_curve_factor(hour_frac):
    xs, ys = ANCHOR_HOURS, ANCHOR_VALS
    if hour_frac <= xs[0]:
        return ys[0]
    if hour_frac >= xs[-1]:
        return ys[-1]
    for i in range(len(xs) - 1):
        if xs[i] <= hour_frac <= xs[i + 1]:
            t = (hour_frac - xs[i]) / (xs[i + 1] - xs[i])
            return ys[i] + t * (ys[i + 1] - ys[i])
    return ys[-1]


def solar_dip(hour_frac, capacity_scaled):
    import math

    exposure = math.exp(-((hour_frac - SOLAR_PEAK_HOUR) ** 2) / (2 * SOLAR_SIGMA_HOURS ** 2))
    return capacity_scaled * SOLAR_PEAK_FRACTION * exposure


def build_location_meta(crosswalk_rows):
    """Deterministic per-location capacity and fabricated meter label, keyed
    by generic_id so it does not depend on either file's column order."""
    import random

    rng = random.Random(CAPACITY_SEED)
    meta = {}
    for idx, row in enumerate(crosswalk_rows):
        capacity_interval_peak = round(rng.uniform(40, 220) / 5) * 5
        meta[row["generic_id"]] = {
            "location_type": row["location_type"],
            "capacity_interval_peak": capacity_interval_peak,
            "meter_label": f"Meter {4001 + idx}",
        }
    return meta


def location_value(meta_entry, dt_start, dt_end, resolution_minutes, rng):
    mid = dt_start + (dt_end - dt_start) / 2
    hour_frac = mid.hour + mid.minute / 60 + mid.second / 3600
    factor = day_curve_factor(hour_frac)
    if mid.weekday() >= 5:
        factor *= WEEKEND_FACTOR

    capacity_scaled = meta_entry["capacity_interval_peak"] * (resolution_minutes / 15.0)
    raw = capacity_scaled * factor

    if meta_entry["location_type"] == "Solar":
        raw -= solar_dip(hour_frac, capacity_scaled)

    raw = max(raw, capacity_scaled * 0.04)
    raw *= 1.0 + rng.uniform(-0.004, 0.004)
    return round(raw, 1)


def month_bounds(report_month):
    year, month = (int(p) for p in report_month.split("-"))
    start = datetime(year, month, 1)
    end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
    return start, end


def generate_timestamps(start, end, minutes):
    stamps = []
    dt = start
    step = timedelta(minutes=minutes)
    while dt < end:
        stamps.append(dt)
        dt += step
    return stamps


def write_workbook(path, ordered_generic_ids, header_lookup, meta, timestamps,
                    resolution_minutes, timestamps_as_text, noise_seed):
    import random

    rng = random.Random(noise_seed)

    wb = Workbook()
    wb.properties.creator = DOC_CREATOR
    wb.properties.lastModifiedBy = DOC_CREATOR
    wb.properties.title = f"Synthetic sample data -- {Path(path).name}"
    wb.properties.company = DOC_COMPANY
    ws = wb.active
    ws.title = SHEET_NAME

    bold = Font(bold=True)
    italic = Font(italic=True)

    # Row 1 — headers
    ws.cell(row=1, column=1, value="Interval Start").font = bold
    ws.cell(row=1, column=2, value="Interval End").font = bold
    for col_offset, generic_id in enumerate(ordered_generic_ids):
        ws.cell(row=1, column=3 + col_offset, value=header_lookup[generic_id]).font = bold
    ws.cell(row=1, column=3 + len(ordered_generic_ids), value=TOTAL_COLUMN_HEADER).font = bold

    # Row 2 — unit/description row
    for col in range(3, 3 + len(ordered_generic_ids) + 1):
        ws.cell(row=2, column=col, value=UNIT_LABEL).font = italic

    # Data rows
    first_data_row = 3
    col_sums = [0.0] * len(ordered_generic_ids)
    for row_offset, dt_start in enumerate(timestamps):
        row = first_data_row + row_offset
        dt_end = dt_start + timedelta(minutes=resolution_minutes)

        if timestamps_as_text:
            ws.cell(row=row, column=1, value=dt_start.strftime(INTERVAL_TIMESTAMP_FORMAT))
            ws.cell(row=row, column=2, value=dt_end.strftime(INTERVAL_TIMESTAMP_FORMAT))
        else:
            c1 = ws.cell(row=row, column=1, value=dt_start)
            c1.number_format = "yyyy-mm-dd hh:mm"
            c2 = ws.cell(row=row, column=2, value=dt_end)
            c2.number_format = "yyyy-mm-dd hh:mm"

        row_total = 0.0
        for col_offset, generic_id in enumerate(ordered_generic_ids):
            value = location_value(meta[generic_id], dt_start, dt_end, resolution_minutes, rng)
            ws.cell(row=row, column=3 + col_offset, value=value).number_format = "#,##0.0"
            col_sums[col_offset] += value
            row_total += value

        row_total = round(row_total, 1)
        ws.cell(row=row, column=3 + len(ordered_generic_ids), value=row_total).number_format = "#,##0.0"

    # Total row
    total_row = first_data_row + len(timestamps)
    ws.cell(row=total_row, column=1, value="Total").font = bold
    grand_total = 0.0
    for col_offset in range(len(ordered_generic_ids)):
        s = round(col_sums[col_offset], 1)
        cell = ws.cell(row=total_row, column=3 + col_offset, value=s)
        cell.font = bold
        cell.number_format = "#,##0.0"
        grand_total += s
    grand_total = round(grand_total, 1)
    gt_cell = ws.cell(row=total_row, column=3 + len(ordered_generic_ids), value=grand_total)
    gt_cell.font = bold
    gt_cell.number_format = "#,##0.0"

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    wb.save(path)
    return {
        "rows": len(timestamps),
        "first_ts": timestamps[0],
        "last_ts": timestamps[-1],
        "grand_total": grand_total,
    }


def main():
    config = load_config()
    crosswalk_rows = load_crosswalk()
    meta = build_location_meta(crosswalk_rows)

    report_month = config["report_month"]
    month_start, month_end = month_bounds(report_month)
    days_in_month = (month_end - month_start).days

    interval_minutes = config["expected_interval_minutes"]
    hourly_minutes = config["hourly_interval_minutes"]

    interval_ts = generate_timestamps(month_start, month_end, interval_minutes)
    hourly_ts = generate_timestamps(month_start, month_end, hourly_minutes)

    interval_order = sorted(crosswalk_rows, key=lambda r: r["interval_position"])
    hourly_order = sorted(crosswalk_rows, key=lambda r: r["hourly_position"])

    interval_ids = [r["generic_id"] for r in interval_order]
    hourly_ids = [r["generic_id"] for r in hourly_order]

    # Interval file uses substation names; hourly file uses fabricated meter
    # labels in a different order -- same as the real vendor exports, which
    # name the same location differently in each file.
    interval_headers = {gid: gid for gid in interval_ids}
    hourly_headers = {gid: meta[gid]["meter_label"] for gid in hourly_ids}

    DATA_DIR.mkdir(exist_ok=True)
    interval_path = DATA_DIR / f"coop_15min_{report_month}_SAMPLE.xlsx"
    hourly_path = DATA_DIR / f"coop_hourly_{report_month}_SAMPLE.xlsx"

    interval_summary = write_workbook(
        interval_path, interval_ids, interval_headers, meta, interval_ts,
        interval_minutes, timestamps_as_text=True, noise_seed=INTERVAL_NOISE_SEED,
    )
    hourly_summary = write_workbook(
        hourly_path, hourly_ids, hourly_headers, meta, hourly_ts,
        hourly_minutes, timestamps_as_text=False, noise_seed=HOURLY_NOISE_SEED,
    )

    print("=== Substation Report Validator — Step 0 sample data generator ===")
    print(f"Report month: {report_month} ({days_in_month} days)")
    print(f"timezone (config): mode={config['timezone']['mode']!r}, zone={config['timezone']['zone']!r}")
    print(f"Interval timestamp format (config, confirmed Aug 12): {INTERVAL_TIMESTAMP_FORMAT!r}")
    print(f"Timestamp column meaning (config, confirmed Aug 12): {TIMESTAMP_COLUMN_MEANING}")
    print(f"Value unit (config, confirmed Aug 11): {config['units']['interval_value_unit']} per interval, "
          f"aggregation_method={config['units']['aggregation_method']}")
    print(f"Random seeds: capacity={CAPACITY_SEED}, interval noise={INTERVAL_NOISE_SEED}, "
          f"hourly noise={HOURLY_NOISE_SEED} (fixed -> reruns are deterministic)")
    print()
    print(f"Wrote {interval_path.relative_to(REPO_ROOT)}")
    print(f"  rows: {interval_summary['rows']}, "
          f"first: {interval_summary['first_ts']}, last: {interval_summary['last_ts']}, "
          f"grand total: {interval_summary['grand_total']:,.1f} kWh")
    print(f"  column order (interval_position): {interval_ids}")
    print()
    print(f"Wrote {hourly_path.relative_to(REPO_ROOT)}")
    print(f"  rows: {hourly_summary['rows']}, "
          f"first: {hourly_summary['first_ts']}, last: {hourly_summary['last_ts']}, "
          f"grand total: {hourly_summary['grand_total']:,.1f} kWh")
    print(f"  column order (hourly_position): {hourly_ids}")
    print()
    diff = interval_summary["grand_total"] - hourly_summary["grand_total"]
    pct = 100 * diff / interval_summary["grand_total"]
    print(f"Grand-total difference (interval - hourly): {diff:,.1f} kWh ({pct:.3f}%) "
          f"-- expected to be small and non-zero, driven by the Sub H/O/P solar dip "
          f"sampled at two different resolutions plus tiny per-meter noise.")


if __name__ == "__main__":
    main()
